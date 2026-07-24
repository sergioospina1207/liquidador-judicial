"""
Generador de Excel con formato completo usando openpyxl
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
import io

AZ_OSC='1F3864'; AZ_MED='2563A8'; AZ_CLR='BDD7EE'
VD_OSC='1E5631'; VD_CLR='C6EFCE'
GR_CLR='F2F2F2'; AMA='FFF2CC'; NAR='FCE4D6'
RJ='7F0000'; BL='FFFFFF'

def fill(c): return PatternFill('solid', fgColor=c)
def fnt(bold=False, color='000000', size=8, italic=False):
    return Font(name='Arial Narrow', bold=bold, color=color, size=size, italic=italic)
def aln(h='left', v='center', wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def brd(style='thin', color='BFBFBF'):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)

THIN = brd(); MED = brd('medium', '595959')
NUM = '#,##0'; DEC4 = '0.0000'; DEC2 = '0.00'

def sc(ws, r, c, v, bold=False, bg=None, fg='000000', size=8,
       h='left', wrap=False, nf=None, it=False, b=None):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = fnt(bold=bold, color=fg, size=size, italic=it)
    cell.alignment = aln(h=h, wrap=wrap)
    if bg: cell.fill = fill(bg)
    if nf: cell.number_format = nf
    if b:  cell.border = b
    return cell

def fr(ws, r, c1, c2, bg, b=None):
    for c in range(c1, c2+1):
        ws.cell(row=r, column=c).fill = fill(bg)
        if b: ws.cell(row=r, column=c).border = b

def set_print(ws):
    ws.page_setup.paperSize = 8
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins = PageMargins(left=0.25, right=0.25, top=0.35, bottom=0.35,
                                   header=0.2, footer=0.2)
    ws.print_options.horizontalCentered = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True

def fmt_fecha(s):
    if not s: return '—'
    s = str(s)
    if len(s) == 10 and s[4] == '-':
        return s[8:]+'/'+s[5:7]+'/'+s[:4]
    return s

def generar_excel(data: dict) -> bytes:
    cliente   = data.get('cliente', {})
    col_hdrs  = data.get('colHeaders', [])
    filas     = data.get('filas', [])
    totales   = data.get('totales', {})
    resumen   = data.get('resumen', {})
    tramos    = data.get('tramosData', [])
    intParams = data.get('intParams', {})

    N = len(col_hdrs)
    TOTAL_COLS = 2 + N + 7  # Período+Días+N prest+BRUTO+IBC+Ded+NETO+IpcIni+Factor+nIdx

    wb = Workbook()
    ws1 = wb.active; ws1.title = 'Resumen Liquidación'
    ws2 = wb.create_sheet('Intereses Moratorios')

    # ══════════════════════════════════════════════════════
    # HOJA 1: RESUMEN
    # ══════════════════════════════════════════════════════
    cw = [22, 5] + [10]*N + [12, 9, 9, 11, 6, 6, 13]
    for i, w in enumerate(cw, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w
    set_print(ws1)
    ws1.print_title_rows = '1:8'

    # R1 Título
    last_col = get_column_letter(TOTAL_COLS)
    ws1.merge_cells(f'A1:{last_col}1')
    sc(ws1,1,1,'LIQUIDACIÓN DE PRESTACIONES SOCIALES SOBRE BONIFICACIÓN JUDICIAL',
       bold=True,bg=AZ_OSC,fg=BL,size=11,h='center',b=MED)
    ws1.row_dimensions[1].height = 20
    ws1.row_dimensions[2].height = 4

    # R3-5 Info cliente
    info = [
        ('NIP:', cliente.get('nip','—'), 'Nombre:', cliente.get('nombre','—'),
         'Ejecutoria:', fmt_fecha(cliente.get('ejecutoria',''))),
        ('Entidad:', cliente.get('entidad','—'), 'Cargo(s):', cliente.get('cargo','—'),
         'Turno Pago:', fmt_fecha(cliente.get('turno',''))),
        ('Período:', cliente.get('periodo','—'), 'IPC Ejecutoria:', cliente.get('ipc','—'), '', ''),
    ]
    for ri, (l1,v1,l2,v2,l3,v3) in enumerate(info, 3):
        ws1.row_dimensions[ri].height = 13
        sc(ws1,ri,1,l1,bold=True,bg=GR_CLR,size=8,b=THIN)
        ws1.merge_cells(f'B{ri}:D{ri}')
        sc(ws1,ri,2,v1,bold=bool(v1),size=8)
        sc(ws1,ri,5,l2,bold=True,bg=GR_CLR,size=8,b=THIN)
        ws1.merge_cells(f'F{ri}:H{ri}')
        sc(ws1,ri,6,v2,bold=bool(v2),size=8)
        sc(ws1,ri,9,l3,bold=True,bg=GR_CLR,size=8,b=THIN)
        ws1.merge_cells(f'J{ri}:L{ri}')
        sc(ws1,ri,10,v3,size=8)

    ws1.row_dimensions[6].height = 4

    # R7 Encabezado grupos
    LP = 2 + N  # última columna de prestaciones
    ws1.row_dimensions[7].height = 16
    sc(ws1,7,1,'Período',bold=True,bg=AZ_OSC,fg=BL,h='center',b=THIN)
    sc(ws1,7,2,'Días',bold=True,bg=AZ_OSC,fg=BL,h='center',b=THIN)
    ws1.merge_cells(f'C7:{get_column_letter(LP)}7')
    sc(ws1,7,3,'PRESTACIONES SOCIALES',bold=True,bg=AZ_MED,fg=BL,h='center',b=THIN)
    fr(ws1,7,4,LP,AZ_MED,THIN)
    extras = [('BRUTO',RJ),('IBC','7F3F00'),('Deduc.\n9%',RJ),
              ('NETO',VD_OSC),('IPC\nIni',AZ_OSC),('Factor',AZ_OSC),('NETO\nINDEXADO',VD_OSC)]
    for i,(lbl,bgc) in enumerate(extras, LP+1):
        sc(ws1,7,i,lbl,bold=True,bg=bgc,fg=BL,size=7,h='center',wrap=True,b=THIN)

    # R8 Nombres prestaciones
    ws1.row_dimensions[8].height = 26
    fr(ws1,8,1,2,AZ_OSC,THIN); fr(ws1,8,LP+1,TOTAL_COLS,AZ_OSC,THIN)
    for ci,h in enumerate(col_hdrs, 3):
        sc(ws1,8,ci,h,bold=True,bg=AZ_MED,fg=BL,size=6,h='center',wrap=True,b=THIN)

    # Filas datos
    cr = 9; alt = False
    for fila in filas:
        if fila['tipo'] == 'sep':
            ws1.merge_cells(f'A{cr}:{last_col}{cr}')
            sc(ws1,cr,1,'── '+fila['label']+' ──',bold=True,bg=AZ_CLR,fg=AZ_OSC,it=True,b=THIN)
            ws1.row_dimensions[cr].height = 12; cr += 1; alt = False
        else:
            bg2 = GR_CLR if alt else BL
            ws1.row_dimensions[cr].height = 11
            sc(ws1,cr,1,fila['periodo'],bg=bg2,size=7,b=THIN)
            sc(ws1,cr,2,fila['dias'],bg=bg2,size=7,h='center',nf='0',b=THIN)
            for ci,v in enumerate(fila['cols'], 3):
                hv = v and v != 0
                sc(ws1,cr,ci,v if hv else '',bg=AMA if hv else bg2,
                   size=7,h='right',nf=NUM if hv else None,b=THIN)
            sc(ws1,cr,LP+1,fila['bruto'],bold=True,bg=NAR,size=7,h='right',nf=NUM,b=THIN)
            sc(ws1,cr,LP+2,fila['ibc'] or '',bg=bg2,size=7,h='right',
               nf=NUM if fila['ibc'] else None,b=THIN)
            sc(ws1,cr,LP+3,fila['ded'] or '',bg=NAR,size=7,h='right',
               nf=NUM if fila['ded'] else None,b=THIN)
            sc(ws1,cr,LP+4,fila['neto'],bold=True,bg=VD_CLR,fg=VD_OSC,size=7,h='right',nf=NUM,b=THIN)
            sc(ws1,cr,LP+5,fila['ipcIni'] or '',bg=bg2,size=7,h='center',nf=DEC2,b=THIN)
            sc(ws1,cr,LP+6,fila['factor'],bg=bg2,size=7,h='center',nf=DEC4,b=THIN)
            sc(ws1,cr,LP+7,fila['nIdx'],bold=True,bg=VD_CLR,fg=VD_OSC,size=7,h='right',nf=NUM,b=THIN)
            alt = not alt; cr += 1

    # TOTALES
    ws1.row_dimensions[cr].height = 14
    sc(ws1,cr,1,'TOTALES',bold=True,bg=AZ_OSC,fg=BL,size=8,b=MED)
    fr(ws1,cr,2,2,AZ_OSC,MED)
    for ci,v in enumerate(totales.get('cols',[]), 3):
        sc(ws1,cr,ci,v or '',bold=True,bg=AZ_OSC,fg=BL,size=7,h='right',
           nf=NUM if v else None,b=MED)
    for i2,v in enumerate([totales.get('bruto',0),totales.get('ibc',0),
                            totales.get('ded',0),totales.get('neto',0)], LP+1):
        sc(ws1,cr,i2,v,bold=True,bg=AZ_OSC,fg=BL,size=7,h='right',nf=NUM,b=MED)
    fr(ws1,cr,LP+5,LP+6,AZ_OSC,MED)
    sc(ws1,cr,LP+7,totales.get('nIdx',0),bold=True,bg=VD_OSC,fg=BL,size=7,h='right',nf=NUM,b=MED)

    # DIFERENCIA
    cr += 1; ws1.row_dimensions[cr].height = 12
    ws1.merge_cells(f'A{cr}:{get_column_letter(TOTAL_COLS-1)}{cr}')
    sc(ws1,cr,1,'DIFERENCIA POR INDEXACIÓN (Indexado — Neto):',bold=True,bg=VD_OSC,fg=BL,size=8,h='right')
    dif = totales.get('nIdx',0) - totales.get('neto',0)
    sc(ws1,cr,TOTAL_COLS,dif,bold=True,bg=VD_OSC,fg=BL,size=8,h='right',nf=NUM,b=THIN)

    # RESUMEN cuadro derecho
    cr += 2; ws1.row_dimensions[cr].height = 14
    RC1 = LP+1
    ws1.merge_cells(f'{get_column_letter(RC1)}{cr}:{last_col}{cr}')
    sc(ws1,cr,RC1,'RESUMEN',bold=True,bg=AZ_OSC,fg=BL,size=9,h='center')
    fr(ws1,cr,RC1+1,TOTAL_COLS,AZ_OSC,MED)

    res = resumen
    if res.get('modo') == 'resolucion':
        res_rows = [
            ('NETO Indexado', res.get('netoIdx',0), GR_CLR, False),
            (f"Pensión Empleador 12% (base: ${res.get('baseRes',0):,})", res.get('pension',0), GR_CLR, False),
            ('Salud Empleador 8.5%', res.get('salud',0), GR_CLR, False),
            ('TOTAL CON APORTES', res.get('totalAportes',0), AZ_CLR, True),
            (f"Honorarios {res.get('honPct',30)}%", res.get('honV',0), NAR, False),
            *([('IVA 19% sobre honorarios', res.get('honIvaV',0), NAR, False)]
              if res.get('honIVA')=='noasume' else []),
            ('VALOR FINAL CLIENTE', res.get('valorFinal',0), VD_OSC, True),
        ]
    else:
        res_rows = [
            ('NETO Indexado', res.get('netoIdx',0), GR_CLR, False),
            ('Total Intereses Moratorios', res.get('intereses',0), GR_CLR, False),
            ('TOTAL BASE (Indexado + Intereses)', res.get('totalBase',0), AZ_CLR, True),
            (f"Descuento {res.get('pctDesc',15.5)}%", res.get('descV',0), NAR, False),
            ('VALOR COMPRA NETO', res.get('valorCompra',0), VD_CLR, True),
            (f"Honorarios {res.get('honPct',30)}%", res.get('honV',0), NAR, False),
            *([('IVA 19% sobre honorarios', res.get('honIvaV',0), NAR, False)]
              if res.get('honIVA')=='noasume' else []),
            ('VALOR FINAL CLIENTE', res.get('valorFinal',0), VD_OSC, True),
        ]

    for lbl, val, bgr, bld in res_rows:
        cr += 1; ws1.row_dimensions[cr].height = 13
        ws1.merge_cells(f'{get_column_letter(RC1)}{cr}:{get_column_letter(TOTAL_COLS-1)}{cr}')
        fgr = BL if bgr == VD_OSC else '000000'
        sc(ws1,cr,RC1,lbl,bold=bld,bg=bgr,fg=fgr,size=8,b=THIN)
        fr(ws1,cr,RC1+1,TOTAL_COLS-1,bgr,THIN)
        sc(ws1,cr,TOTAL_COLS,val,bold=bld,bg=bgr,fg=fgr,size=8,h='right',nf=NUM,b=THIN)

    ws1.freeze_panes = 'A9'

    # ══════════════════════════════════════════════════════
    # HOJA 2: INTERESES
    # ══════════════════════════════════════════════════════
    cw2 = [4,12,12,5,7,9,13,14,10]
    for i,w in enumerate(cw2,1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    set_print(ws2)
    ws2.print_title_rows = '1:1'

    ws2.merge_cells('A1:I1')
    sc(ws2,1,1,'LIQUIDACIÓN DE INTERESES MORATORIOS',bold=True,bg=AZ_OSC,fg=BL,size=11,h='center',b=MED)
    ws2.row_dimensions[1].height = 20
    ws2.row_dimensions[2].height = 4

    ws2.row_dimensions[3].height = 13
    sc(ws2,3,1,'NIP:',bold=True,bg=GR_CLR,size=8,b=THIN)
    sc(ws2,3,2,cliente.get('nip',''),bold=True,size=8)
    sc(ws2,3,4,'Nombre:',bold=True,bg=GR_CLR,size=8,b=THIN)
    ws2.merge_cells('E3:I3')
    sc(ws2,3,5,cliente.get('nombre',''),bold=True,size=8)
    ws2.row_dimensions[4].height = 4

    ip = intParams
    cr2 = 5
    params = [
        ('Capital nominal sentencia:', ip.get('capital',0)),
        ('Fecha inicio (ejecutoria):', fmt_fecha(ip.get('fechaIni',''))),
        ('Fecha radicación:', fmt_fecha(ip.get('fechaRad',''))),
        ('Fecha corte:', fmt_fecha(ip.get('fechaFin',''))),
        ('Meses tasa DTF:', ip.get('mesesDTF',10)),
        ('Factor usura:', ip.get('factorUsura',1.5)),
    ]
    for lbl,val in params:
        ws2.row_dimensions[cr2].height = 12
        ws2.merge_cells(f'A{cr2}:C{cr2}')
        sc(ws2,cr2,1,lbl,bold=True,bg=GR_CLR,size=8,b=THIN)
        ws2.merge_cells(f'D{cr2}:F{cr2}')
        sc(ws2,cr2,4,val,bold=True,size=8,nf=NUM if isinstance(val,(int,float)) else None)
        cr2 += 1

    cr2 += 1; ws2.row_dimensions[cr2].height = 4; cr2 += 1

    ws2.row_dimensions[cr2].height = 14
    for ci,h in enumerate(['#','Desde','Hasta','Días','Etapa','Tasa EA %',
                            'Tasa Período %','Interés tramo','Estado'], 1):
        sc(ws2,cr2,ci,h,bold=True,bg=AZ_OSC,fg=BL,size=8,h='center',b=THIN)
    cr2 += 1

    alt2 = False
    for t in tramos:
        ws2.row_dimensions[cr2].height = 11
        est = str(t.get('estado','Activo'))
        bg3 = 'FDECEA' if 'Suspend' in est else (AMA if 'Reactivado' in est else (GR_CLR if alt2 else BL))
        sc(ws2,cr2,1,t.get('n',''),bg=bg3,size=7,h='center',b=THIN)
        sc(ws2,cr2,2,str(t.get('desde','')),bg=bg3,size=7,h='center',b=THIN)
        sc(ws2,cr2,3,str(t.get('hasta','')),bg=bg3,size=7,h='center',b=THIN)
        sc(ws2,cr2,4,t.get('dias',0),bg=bg3,size=7,h='center',nf='0',b=THIN)
        sc(ws2,cr2,5,t.get('etapa',''),bg=bg3,size=7,h='center',b=THIN)
        sc(ws2,cr2,6,t.get('tasaEA',0),bg=bg3,size=7,h='right',nf=DEC2,b=THIN)
        sc(ws2,cr2,7,t.get('tasaPer',0),bg=bg3,size=7,h='right',nf=DEC4,b=THIN)
        sc(ws2,cr2,8,t.get('interes',0),bold=True,bg=bg3,size=7,h='right',nf=NUM,b=THIN)
        col_e = AZ_CLR if est=='Activo' else ('FDECEA' if 'Suspend' in est else AMA)
        sc(ws2,cr2,9,est,bg=col_e,size=7,h='center',b=THIN)
        alt2 = not alt2; cr2 += 1

    # Total
    ws2.row_dimensions[cr2].height = 13
    ws2.merge_cells(f'A{cr2}:C{cr2}')
    sc(ws2,cr2,1,'TOTAL',bold=True,bg=AZ_OSC,fg=BL,size=8,b=MED)
    fr(ws2,cr2,2,3,AZ_OSC,MED)
    sc(ws2,cr2,4,sum(t.get('dias',0) for t in tramos),bold=True,bg=AZ_OSC,fg=BL,size=7,h='center',nf='0',b=MED)
    fr(ws2,cr2,5,7,AZ_OSC,MED)
    sc(ws2,cr2,8,ip.get('totalIntereses',0),bold=True,bg=AZ_OSC,fg=BL,size=7,h='right',nf=NUM,b=MED)
    fr(ws2,cr2,9,9,AZ_OSC,MED); cr2 += 2

    # Resumen intereses
    ws2.merge_cells(f'A{cr2}:I{cr2}')
    sc(ws2,cr2,1,'RESUMEN',bold=True,bg=AZ_OSC,fg=BL,size=9,h='center')
    fr(ws2,cr2,2,9,AZ_OSC,MED); ws2.row_dimensions[cr2].height = 14

    cap    = ip.get('capital',0) or 0
    int_t  = ip.get('totalIntereses',0) or 0
    total_b = cap + int_t
    desc_p = (ip.get('descuento',15.5) or 15.5)/100
    desc_v  = total_b * desc_p
    v_comp  = total_b - desc_v

    res = data.get('resumen',{})
    if res.get('modo')=='resolucion':
        res2_rows=[
            ('NETO Indexado (prestaciones)', res.get('netoIdx',0), GR_CLR, False),
            (f"Pensión Empleador 12% (base bonif. ${res.get('baseRes',0):,})", res.get('pension',0), GR_CLR, False),
            ('Salud Empleador 8.5%', res.get('salud',0), GR_CLR, False),
            ('Total con Aportes', res.get('totalAportes',0), AZ_CLR, True),
            ('Intereses moratorios', res.get('intereses',0), GR_CLR, False),
            ('TOTAL A PAGAR', res.get('totalFin',0), AZ_OSC, True),
            (f"Honorarios {res.get('honPct',30)}%"+(' + IVA 19%' if res.get('honIVA')=='noasume' else ' (IVA incluido)'), res.get('honV',0)+res.get('honIvaV',0), NAR, False),
            ('VALOR FINAL CLIENTE', res.get('valorFinal',0), VD_OSC, True),
        ]
    else:
        res2_rows=[
            ('NETO Indexado (prestaciones)', res.get('netoIdx',0), GR_CLR, False),
            ('Intereses moratorios', res.get('intereses',0), GR_CLR, False),
            ('TOTAL A PAGAR', res.get('totalBase',0), AZ_CLR, True),
            (f"Descuento {res.get('pctDesc',15.5)}%", res.get('descV',0), NAR, False),
            ('VALOR COMPRA NETO', res.get('valorCompra',0), VD_CLR, True),
            (f"Honorarios {res.get('honPct',30)}%"+(' + IVA 19%' if res.get('honIVA')=='noasume' else ' (IVA incluido)'), res.get('honV',0)+res.get('honIvaV',0), NAR, False),
            ('VALOR FINAL CLIENTE', res.get('valorFinal',0), VD_OSC, True),
        ]
    for lbl,val,bgr,bld in res2_rows:
        cr2 += 1; ws2.row_dimensions[cr2].height = 13
        ws2.merge_cells(f'A{cr2}:G{cr2}')
        fgr = BL if bgr==VD_OSC else '000000'
        sc(ws2,cr2,1,lbl,bold=bld,bg=bgr,fg=fgr,size=8,b=THIN)
        fr(ws2,cr2,2,7,bgr,THIN)
        ws2.merge_cells(f'H{cr2}:I{cr2}')
        sc(ws2,cr2,8,val,bold=bld,bg=bgr,fg=fgr,size=8,h='right',nf=NUM if val else None,b=THIN)

    ws2.freeze_panes = 'A13'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
